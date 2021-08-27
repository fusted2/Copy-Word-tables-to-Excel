import os, logging, locale
import openpyxl, docx # third-party package --pip install--
from openpyxl.styles.borders import Border, Side # to apply border
from openpyxl.styles import Font # to format Excell cells

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
# logging.disable(logging.DEBUG, logging.CRITICAL)

# set thousand & decimal separate:
locale.setlocale(locale.LC_NUMERIC, "en_DK.UTF-8") # 1,234,567.999
# locale.setlocale(locale.LC_NUMERIC, "en_US.UTF-8") # 1.234.567,999

logging.critical("Running.......\n")

# change directory to folder containing the Word file
os.chdir(r'D:\example')

# load existed Excel and Word files
# sometimes the Excel file is already having some functions (sum, countif,...), so I don't want to load a blank Excel file.
wb = openpyxl.load_workbook('blankExcelReport.xlsx')
doc = docx.Document('wordReport.docx')

# call the sheet you want to copy to
sheet1 = wb['Sheet1']
# sheet2 = wb['Sheet2']
# sheet3 = wb['Sheet3']
# sheet4 = wb['Sheet4']
# sheet5 = wb['Sheet5']

# set border style for Excel cells
thin_border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

# I want to copy tables from Word to Excel in consecutive order and separate them by 3 blank rows.
# [1] I want all tables copied into only 1 sheet
# choose only [1] or [2] at a time
rowNumEx = 0
activeSheet = wb['Sheet1']

# loop through all tables in the Word file --or-- specify range
for t in range(len(doc.tables)):
    # # (optional) skip some tables --or-- only some tables
    # if t in [0,1,2,3,4,5,9,
    #         13,16,17,18,19,
    #         20,21,22,23,24,25,
    #         87]:
    #     continue
    
    # # [2] (optional) split copied tables into multiple sheets
    # if t <= 8:
    #     activeSheet = sheet1
    # elif t <= 12:
    #     activeSheet = sheet2
    # elif t <= 15:
    #     activeSheet = sheet3
    # elif t <= 79:
    #     activeSheet = sheet4
    # else:
    #     activeSheet = sheet5
    
    # # [2] (optional) reset at A1 whenever moving to new sheet -- recheck with skipped tables above
    # if t in [6,10,14,26,80]:
    #     rowNumEx = 0
    
    # assign the table need to be copied
    tableDoc = doc.tables[t]
    
    # get the size of the table
    rowNumDoc = len(tableDoc.rows)
    colNumDoc = len(tableDoc.columns)

    # create a list ._tc to check merged cells in Word
    checkMerge = []
    
    # loop through all cells in the table
    for i in range(rowNumDoc):
        for j in range(colNumDoc):
            # Excel cell starts at (1, 1)
            excelCell = activeSheet.cell(row = i+1+rowNumEx, column = j+1)
            
            # Word cell starts at (0, 0) = top left corner of table
            # Convert number from Word to number in Excel
            if tableDoc.cell(i, j)._tc not in checkMerge:
                try:
                    excelCell.value = locale.atoi(tableDoc.cell(i, j).text.replace('(', '-').replace(')', '')) # integer; (4123) = -4123
                except ValueError:
                    try:
                        excelCell.value = locale.atof(tableDoc.cell(i, j).text.replace('(', '-').replace(')', '')) # float; (4.123) = -4.123
                    except ValueError:
                        excelCell.value = tableDoc.cell(i, j).text # string
            
            # add ._tc to list
            checkMerge.append(tableDoc.cell(i, j)._tc)
            
            # (optional) format Excel cells
            excelCell.border = thin_border # apply cell border
            if tableDoc.cell(i, j).text: # exclude empty cells in Word
                checkBold = tableDoc.cell(i, j).paragraphs[0].runs[0].bold # check format in Word: True/False/None
                checkItalic = tableDoc.cell(i, j).paragraphs[0].runs[0].italic # check format in Word: True/False/None
                checkUnderline = tableDoc.cell(i, j).paragraphs[0].runs[0].underline # check format in Word: True/False/None
            
                # for --underline--: Value must be one of {'single', 'singleAccounting', 'doubleAccounting', 'double'}
                excelCell.font = Font(bold = checkBold, italic = checkItalic, underline = ('single' if checkUnderline else None))
        
        # # (optional) just in case we need to know the position of the table copied from Word
        # tableNote = activeSheet.cell(row = i+1+rowNumEx, column = 26) # column Z
        # tableNote.value = 'This is table %s' % (t)
    
    # mentioned above: consecutive order and separate by 3 rows
    rowNumEx += rowNumDoc + 3

# save a new Excel file just in case the code messes everything up
wb.save('newExcelReport.xlsx')

# Done message
logging.critical("""
**************
*            *
*    Done    *
*            *
**************
""")